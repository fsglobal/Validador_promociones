import os
import re
import sys
import json
from datetime import datetime
from io import StringIO, BytesIO
import xml.etree.ElementTree as ET

import pandas as pd
from flask import Flask, render_template, request, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ============================================================
# RUTAS BASE Y CONFIGURACIÓN INICIAL
# ============================================================
BASE_PATH = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
MODULOS_PATH = os.path.join(BASE_PATH, "modulos")
EXCEL_PATH = os.path.join(BASE_PATH, "Excel")
EXPORT_PATH = os.path.join(BASE_PATH, "Export")
LOG_PATH = os.path.join(BASE_PATH, "logs")

os.makedirs(LOG_PATH, exist_ok=True)
os.makedirs(EXCEL_PATH, exist_ok=True)
os.makedirs(EXPORT_PATH, exist_ok=True)
# Limpieza automática al iniciar el servidor
for carpeta in [EXCEL_PATH, EXPORT_PATH]:
    try:
        for archivo in os.listdir(carpeta):
            ruta = os.path.join(carpeta, archivo)
            if os.path.isfile(ruta):
                os.remove(ruta)
    except Exception:
        pass

if MODULOS_PATH not in sys.path:
    sys.path.append(MODULOS_PATH)

# ============================================================
# IMPORTACIÓN DE MÓDULOS DEL PROYECTO
# ============================================================
from validador import (
    leer_hoja_eventos,
    leer_hoja_completar,
    leer_hoja_imput,
    ejecutar_flujo_tradicional,
    ejecutar_flujo_eventos,
    validar_promocion_tradicional,
    validar_promocion_eventos,
    validar_promocion_completar,
    normalizar_local,
    normalizar_texto,
    parsear_promos,
    convertir_txt_a_xml_con_root,
)

try:
    from parser_listas_export import parsear_listas_productos_export
except Exception:
    def parsear_listas_productos_export(_ruta):
        return {}

try:
    from gestor import registrar_rutas_gestor
except Exception:
    def registrar_rutas_gestor(_app):
        return None

try:
    from repositorio import registrar_rutas_repositorio
except Exception:
    def registrar_rutas_repositorio(_app):
        return None

try:
    from consultor.reglas_consultor import REGLAS_CONSULTOR
    from consultor.consultor_carga import construir_consulta
except Exception:
    try:
        from reglas_consultor import REGLAS_CONSULTOR
        from consultor_carga import construir_consulta
    except Exception:
        REGLAS_CONSULTOR = {
            "catalogos": {
                "modalidades": [],
                "submodos_club": [],
                "areas_funcionales": [],
                "tipos_descuento": [],
            },
            "textos": {},
            "requiere_mensaje_tipos": [],
        }
        construir_consulta = None

# ============================================================
# CONFIGURACIÓN FLASK
# ============================================================
app = Flask(
    __name__,
    template_folder=os.path.join(os.path.dirname(__file__), "templates"),
    static_folder=os.path.join(os.path.dirname(__file__), "static"),
)
app.secret_key = "ClaveUltraSecretaParaMensajesWeb"
registrar_rutas_gestor(app)
registrar_rutas_repositorio(app)

ULTIMO_REPORTE_DESCARGA = {
    "rc": "",
    "tradicional": [],
    "completar": [],
}

# ============================================================
# LOGGING
# ============================================================
def escribir_log(linea):
    archivo = os.path.join(LOG_PATH, f"log_{datetime.now().strftime('%Y-%m-%d')}.txt")
    with open(archivo, "a", encoding="utf-8") as f:
        f.write(f"[{datetime.now().strftime('%H:%M:%S')}] {linea}\n")


# ============================================================
# UTILIDADES GENERALES DE ARCHIVOS
# ============================================================
def limpiar_carpeta(path):
    errores = []
    for f in os.listdir(path):
        try:
            fp = os.path.join(path, f)
            if os.path.isfile(fp):
                os.remove(fp)
        except Exception as e:
            errores.append(str(e))
    return errores


def listar_archivos():
    return sorted(os.listdir(EXCEL_PATH)), sorted(os.listdir(EXPORT_PATH))


# ============================================================
# UTILIDADES DE LIMPIEZA Y FORMATEO
# ============================================================
def _strip_html(texto):
    return re.sub(r"<[^>]+>", "", str(texto or "")).strip()


def _extraer_entre_parentesis(texto, etiqueta):
    patron = rf"{re.escape(etiqueta)}\s*:?\s*\((.*?)\)"
    m = re.search(patron, texto)
    return m.group(1).strip() if m else ""


def _normalizar_lista_valores(valor):
    if not valor or valor == "-":
        return "-"
    partes = [p.strip() for p in valor.split(",") if p.strip()]
    return " - ".join(partes) if partes else "-"


def _formatear_monto_limpio(valor):
    if not valor or valor == "-":
        return "-"
    try:
        num = float(str(valor).replace(",", "."))
        return f"${int(num)}" if num.is_integer() else f"${num:.2f}"
    except Exception:
        return str(valor)


def _formatear_numero_limpio(valor):
    if not valor or valor == "-":
        return "-"
    try:
        num = float(str(valor).replace(",", "."))
        return f"{num:.2f}"
    except Exception:
        return str(valor)


def _detalle_a_tipo_msg(det):
    if isinstance(det, tuple):
        return det[0], det[1]
    return det.get("tipo"), det.get("msg")


def _detalle_plain(det):
    _tipo, _msg = _detalle_a_tipo_msg(det)
    return _strip_html(_msg)


def _es_caso_especial_bycp_3x2(analisis, detalles):
    tipo = str((analisis or {}).get("tipo_promocion", "")).upper()
    area = str((analisis or {}).get("area_responsable", "")).upper()
    if "PACK 3X2" not in tipo or area != "BYCP":
        return False

    plains = [_detalle_plain(d) for d in (detalles or [])]
    tiene_percentage = any("PERCENTAGE" in p.upper() or "PERCENTAGEDISCOUNTAPPLIER" in p.upper() for p in plains)
    tiene_cant3 = any("Cantidad: (3)" in p or "Cantidad (3)" in p or "cada 3" in p.lower() for p in plains)
    tiene_cant1 = any("Cantidad: (1)" in p or "Cantidad (1)" in p for p in plains)
    tiene_100 = any("%: (100" in p or "% nodo export: (1" in p or "percentage>1.0" in p.lower() or "(100.00)" in p for p in plains)
    return tiene_percentage and tiene_cant3 and tiene_cant1 and tiene_100


def _ajustar_detalles_caso_especial_bycp_3x2(detalles):
    plains = [_detalle_plain(d) for d in (detalles or [])]

    lista = "-"
    for p in plains:
        if p.startswith("[LEYENDA] Condición Export"):
            lista = _extraer_entre_parentesis(p, "Lista") or _extraer_entre_parentesis(p, "SKU") or lista
            if lista != "-":
                break
    if lista == "-":
        for p in plains:
            if p.startswith("[LEYENDA] Applier Export"):
                lista = _extraer_entre_parentesis(p, "Lista") or _extraer_entre_parentesis(p, "SKU") or lista
                if lista != "-":
                    break

    porcentaje = "100.00"
    for p in plains:
        if p.startswith("[LEYENDA] Applier Export"):
            porcentaje = (
                _extraer_entre_parentesis(p, "% nodo export")
                or _extraer_entre_parentesis(p, "%")
                or porcentaje
            )
            break

    nuevos = []
    # preservamos solo mensajes neutros/útiles y descartamos la lógica vieja de PACK
    for d in (detalles or []):
        tipo, msg = _detalle_a_tipo_msg(d)
        plain = _strip_html(msg)
        if plain.startswith("[COMPETENCIA]") or plain.startswith("[CONDICIÓN]") or plain.startswith("[APPLIER]"):
            continue
        if plain.startswith("[LEYENDA] Condición Export") or plain.startswith("[LEYENDA] Applier Export"):
            continue
        if plain.startswith("[LEYENDA] Excel → Tipo:") or plain.startswith("[LEYENDA] Excel -> Tipo:"):
            msg = "[LEYENDA] Excel → Tipo: (PACK 3X2) Caso especial BYCP 3X2."
            tipo = "OK"
        nuevos.append((tipo, msg))

    nuevos.append(("OK", "[COMPETENCIA] Caso especial BYCP 3X2: competencia por producto correcta."))
    if lista != "-":
        nuevos.append(("OK", f"[CONDICIÓN] Caso especial BYCP 3X2 válido. Lista: ({lista}) | Cantidad: (3)"))
        nuevos.append(("OK", f"[LEYENDA] Condición Export → Tipo: (LISTA) Lista: ({lista}) Cantidad: (3)"))
        nuevos.append(("OK", f"[APPLIER] Caso especial BYCP 3X2 válido. PercentageDiscountApplier | Lista: ({lista}) | Cantidad: (1) | %: ({porcentaje}) | Strategy: (1)"))
        nuevos.append(("OK", f"[LEYENDA] Applier Export → Tipo: (PERCENTAGE) Lista: ({lista}) Cantidad: (1) % nodo export: ({porcentaje})"))
    else:
        nuevos.append(("OK", "[CONDICIÓN] Caso especial BYCP 3X2 válido. Cantidad: (3)"))
        nuevos.append(("OK", "[APPLIER] Caso especial BYCP 3X2 válido. PercentageDiscountApplier | Cantidad: (1) | %: (100.00) | Strategy: (1)"))
        nuevos.append(("OK", "[LEYENDA] Condición Export → Tipo: (LISTA) Cantidad: (3)"))
        nuevos.append(("OK", "[LEYENDA] Applier Export → Tipo: (PERCENTAGE) Cantidad: (1) % nodo export: (100.00)"))

    return nuevos


def _forzar_analisis_caso_especial_bycp_3x2(analisis):
    a = dict(analisis or {})
    a["tipo_promocion"] = "PACK 3X2"
    a["estado_condicion"] = "Coinciden"
    a["estado_applier"] = "Coinciden"
    a["mensaje_principal"] = "Coinciden"
    a["aviso_principal"] = ""
    if str(a.get("area_responsable", "")).upper() == "BYCP":
        a["resumen_aplicador"] = "Caso especial BYCP 3X2 | % 100 | Cantidad: 1 | Strategy: menor"
    return a


def _formatear_porcentaje_limpio(valor):
    return "-" if not valor or valor == "-" else valor


# ============================================================
# ANÁLISIS DE DETALLES PARA RESUMEN WEB
# ============================================================
def analizar_detalles(detalles):
    mensajes = []
    for d in detalles:
        if isinstance(d, tuple):
            tipo, msg = d
        else:
            tipo, msg = d.get("tipo"), d.get("msg")
        mensajes.append({"tipo": tipo, "msg": msg, "msg_plain": _strip_html(msg)})

    resumen = {
        "estado_id": "No evaluado",
        "area_responsable": "-",
        "estado_facturar": "No evaluado",
        "estado_fechas": "No evaluado",
        "estado_condicion": "No evaluado",
        "estado_applier": "No evaluado",
        "fecha_inicio_ok": None,
        "fecha_fin_ok": None,
        "tipo_promocion": "-",
        "resumen_condicion": "-",
        "resumen_aplicador": "-",
        "restriccion_dias": "-",
        "restriccion_dias_estado": "No evaluado",
        "mensaje_principal": "No coinciden",
        "aviso_principal": "",
    }

    id_items = [x for x in mensajes if x["msg_plain"].startswith("[ID]")]
    fact_items = [x for x in mensajes if x["msg_plain"].startswith("[FACTURAR]")]
    fechas_items = [x for x in mensajes if x["msg_plain"].startswith("[FECHAS]")]
    condicion_items = [x for x in mensajes if x["msg_plain"].startswith("[CONDICIÓN]")]
    applier_items = [x for x in mensajes if x["msg_plain"].startswith("[APPLIER]")]
    leyenda_items = [x for x in mensajes if x["msg_plain"].startswith("[LEYENDA]")]
    area_items = [x for x in mensajes if x["msg_plain"].startswith("[ÁREA]")]
    descuento_items = [x for x in mensajes if x["msg_plain"].startswith("[DESCUENTO]")]
    lista_items = [x for x in mensajes if x["msg_plain"].startswith("[LISTA PRODUCTOS]")]
    locales_items = [x for x in mensajes if x["msg_plain"].startswith("[LOCALES]")]
    dias_items = [x for x in mensajes if x["msg_plain"].startswith("[DÍAS]") or x["msg_plain"].startswith("[DIAS]")]
    msje_items = [x for x in mensajes if x["msg_plain"].startswith("[MSJE]")]

    if area_items:
        m_area = re.search(r"AreaResponsable detectada:\s*\((.*?)\)", area_items[0]["msg_plain"], re.IGNORECASE)
        if m_area:
            resumen["area_responsable"] = m_area.group(1).strip()

    if id_items:
        resumen["estado_id"] = "Coinciden" if all(x["tipo"] == "OK" for x in id_items) else "No coinciden"

    if fact_items:
        if any(x["tipo"] == "ERR" for x in fact_items):
            resumen["estado_facturar"] = "No coinciden"
        elif any(x["tipo"] == "WARN" for x in fact_items):
            resumen["estado_facturar"] = "Advertencia"
        else:
            resumen["estado_facturar"] = "Coinciden"

    inicio_item = next((x for x in fechas_items if x["msg_plain"].startswith("[FECHAS] Fecha Inicio Excel")), None)
    fin_item = next((x for x in fechas_items if x["msg_plain"].startswith("[FECHAS] Fecha Fin Excel")), None)
    inicio_tipo = inicio_item["tipo"] if inicio_item else None
    fin_tipo = fin_item["tipo"] if fin_item else None

    fecha_inicio_excel = ""
    fecha_fin_excel = ""
    if inicio_item:
        resumen["fecha_inicio_ok"] = (inicio_tipo == "OK")
        m = re.search(r"Fecha Inicio Excel \((.*?)\).*?Export \((.*?)\)", inicio_item["msg_plain"], re.IGNORECASE)
        if m:
            fecha_inicio_excel = m.group(1).strip()
    if fin_item:
        resumen["fecha_fin_ok"] = (fin_tipo == "OK")
        m = re.search(r"Fecha Fin Excel \((.*?)\).*?Export \((.*?)\)", fin_item["msg_plain"], re.IGNORECASE)
        if m:
            fecha_fin_excel = m.group(1).strip()

    estado_fechas_base = "No evaluado"
    if inicio_tipo == "OK" and fin_tipo == "OK":
        estado_fechas_base = "OK"
    elif inicio_tipo in {"WARN", "ERR"} and fin_tipo == "OK":
        estado_fechas_base = "Posible Extensión"
    elif fin_tipo == "ERR":
        estado_fechas_base = "No coinciden"
    elif inicio_item or fin_item:
        estado_fechas_base = "No coinciden"

    detalle_fechas = []
    if fecha_inicio_excel:
        detalle_fechas.append(f"Inicio: {fecha_inicio_excel}")
    if fecha_fin_excel:
        detalle_fechas.append(f"Fin: {fecha_fin_excel}")
    resumen["estado_fechas"] = f"{estado_fechas_base} | {' | '.join(detalle_fechas)}" if detalle_fechas and estado_fechas_base != "No evaluado" else estado_fechas_base

    if condicion_items:
        if any(x["tipo"] == "ERR" for x in condicion_items):
            resumen["estado_condicion"] = "No coinciden"
        elif any(x["tipo"] == "WARN" for x in condicion_items):
            resumen["estado_condicion"] = "Advertencia"
        else:
            resumen["estado_condicion"] = "Coinciden"

    applier_sin_sku_explicito = any("no informa sku explícitos" in x["msg_plain"].lower() or "no informa sku explicitos" in x["msg_plain"].lower() for x in applier_items)
    if applier_items:
        if applier_sin_sku_explicito or any(x["tipo"] == "ERR" for x in applier_items):
            resumen["estado_applier"] = "No coinciden"
        elif any(x["tipo"] == "WARN" for x in applier_items):
            resumen["estado_applier"] = "Advertencia"
        else:
            resumen["estado_applier"] = "Coinciden"

    leyenda_excel = next((x for x in leyenda_items if "Excel → Tipo:" in x["msg_plain"]), None)
    leyenda_cond = next((x for x in leyenda_items if "Condición Export →" in x["msg_plain"]), None)
    leyenda_applier = next((x for x in leyenda_items if "Applier Export →" in x["msg_plain"]), None)

    if leyenda_excel:
        tipo = _extraer_entre_parentesis(leyenda_excel["msg_plain"], "Tipo")
        resumen["tipo_promocion"] = tipo if tipo else "-"

    tipo_prom = (resumen["tipo_promocion"] or "").upper()
    es_msje = "MSJE" in tipo_prom or any("MSJE / POPUP" in x["msg_plain"] for x in msje_items)
    es_2da = "2DA" in tipo_prom
    es_pack = bool(re.search(r"\bPACK\b", tipo_prom) or re.search(r"\d+\s*X\s*\d+", tipo_prom))
    es_porcentual = ("PORCENT" in tipo_prom or "%" in tipo_prom)
    es_nominal = ("NOMINAL" in tipo_prom and "PACK NOMINAL" not in tipo_prom)

    if leyenda_cond:
        sku_val = _extraer_entre_parentesis(leyenda_cond["msg_plain"], "SKU")
        lista_val = _extraer_entre_parentesis(leyenda_cond["msg_plain"], "Lista")
        cantidad_cond_val = _extraer_entre_parentesis(leyenda_cond["msg_plain"], "Cantidad")
        sku_fmt = _normalizar_lista_valores(sku_val)
        lista_fmt = lista_val if lista_val and lista_val != "-" else "-"
        partes_cond = []
        if sku_fmt != "-":
            partes_cond.append(f"SKU: {sku_fmt}")
        elif lista_fmt != "-":
            partes_cond.append(f"Lista: {lista_fmt}")
        if es_2da and cantidad_cond_val and cantidad_cond_val not in {"-", "0", "0.0", "0.00"}:
            try:
                q = float(cantidad_cond_val)
                partes_cond.append(f"Cada {int(q) if q.is_integer() else q} unidades")
            except Exception:
                partes_cond.append(f"Cada {cantidad_cond_val} unidades")
        resumen["resumen_condicion"] = " | ".join(partes_cond) if partes_cond else "-"

    if leyenda_applier:
        sku_val = _extraer_entre_parentesis(leyenda_applier["msg_plain"], "SKU")
        lista_val = _extraer_entre_parentesis(leyenda_applier["msg_plain"], "Lista")
        cantidad_val = _extraer_entre_parentesis(leyenda_applier["msg_plain"], "Cantidad")
        porcentaje_val = _extraer_entre_parentesis(leyenda_applier["msg_plain"], "%")
        monto_val = _extraer_entre_parentesis(leyenda_applier["msg_plain"], "Monto")
        monto_export_val = _extraer_entre_parentesis(leyenda_applier["msg_plain"], "Monto export")
        pct_nodo_export = _extraer_entre_parentesis(leyenda_applier["msg_plain"], "% nodo export")
        pct_comercial_excel = _extraer_entre_parentesis(leyenda_applier["msg_plain"], "% comercial Excel")
        dcto_bruto_q = _extraer_entre_parentesis(leyenda_applier["msg_plain"], "Dcto bruto Excel(Q)")
        pvp_pack_excel = _extraer_entre_parentesis(leyenda_applier["msg_plain"], "PVPOfertaPack Excel")
        if not pvp_pack_excel and leyenda_excel:
            pvp_pack_excel = _extraer_entre_parentesis(leyenda_excel["msg_plain"], "PVPOfertaPack")
        unidades_excel = _extraer_entre_parentesis(leyenda_excel["msg_plain"], "Unidades") if leyenda_excel else ""
        sku_fmt = _normalizar_lista_valores(sku_val)
        lista_fmt = _normalizar_lista_valores(lista_val)
        partes = []
        if sku_fmt != "-":
            partes.append(f"SKU: {sku_fmt}")
        elif lista_fmt != "-":
            partes.append(f"Lista: {lista_fmt}")
        if es_msje:
            salida = _extraer_entre_parentesis(leyenda_applier["msg_plain"], "Salida")
            mensaje = _extraer_entre_parentesis(leyenda_applier["msg_plain"], "Mensaje")
            texto = _extraer_entre_parentesis(leyenda_applier["msg_plain"], "Texto")
            if mensaje and mensaje != "-":
                partes.append(f"Mensaje: {mensaje}")
            if salida and salida != "-":
                partes.append(f"Salida: {salida}")
            if texto and texto != "-":
                partes.append(f"Texto: {texto}")
        elif es_2da:
            if cantidad_val and cantidad_val not in {"-", "0", "0.0", "0.00"}:
                try:
                    q = float(cantidad_val)
                    partes.append(f"Cada {int(q) if q.is_integer() else q} unidades")
                except Exception:
                    partes.append(f"Cada {cantidad_val} unidades")
            pct_aplicador_fmt = _formatear_porcentaje_limpio(pct_nodo_export)
            pct_comercial_fmt = _formatear_porcentaje_limpio(pct_comercial_excel)
            dcto_bruto_fmt = _formatear_numero_limpio(dcto_bruto_q)
            if pct_aplicador_fmt != "-":
                partes.append(f"{pct_aplicador_fmt} aplicador")
            if pct_comercial_fmt != "-":
                partes.append(f"{pct_comercial_fmt} comercial")
            if dcto_bruto_fmt != "-":
                partes.append(f"Dcto bruto Q: {dcto_bruto_fmt}")
        elif es_pack:
            monto_fmt = _formatear_monto_limpio(monto_val or monto_export_val)
            pvp_fmt = _formatear_monto_limpio(pvp_pack_excel)
            if cantidad_val and cantidad_val not in {"-", "0", "0.0", "0.00"}:
                try:
                    q = float(cantidad_val)
                    partes.append(f"Cantidad: {int(q) if q.is_integer() else q}")
                except Exception:
                    partes.append(f"Cantidad: {cantidad_val}")
            if monto_fmt != "-":
                partes.append(f"Monto unitario: {monto_fmt}")
            if pvp_fmt != "-" and unidades_excel and unidades_excel != "-":
                try:
                    q_pack = float(unidades_excel)
                    q_txt = int(q_pack) if q_pack.is_integer() else q_pack
                except Exception:
                    q_txt = unidades_excel
                partes.append(f"Pack: {pvp_fmt} / {q_txt}")
        elif es_nominal:
            pvp_fmt = _formatear_monto_limpio(pvp_pack_excel)
            monto_export_fmt = _formatear_monto_limpio(monto_export_val or monto_val)
            dcto_bruto_fmt = _formatear_numero_limpio(dcto_bruto_q)
            if pvp_fmt != "-":
                partes.append(f"PVPOfertaPack: {pvp_fmt}")
            if monto_export_fmt != "-":
                partes.append(f"Monto export: {monto_export_fmt}")
            if dcto_bruto_fmt != "-":
                partes.append(f"Dcto bruto Q: {dcto_bruto_fmt}")
        elif es_porcentual:
            pct_fmt = _formatear_porcentaje_limpio(porcentaje_val or pct_nodo_export)
            if pct_fmt != "-":
                partes.append(f"%: {pct_fmt}")
        else:
            if cantidad_val and cantidad_val not in {"-", "0", "0.0", "0.00"}:
                try:
                    q = float(cantidad_val)
                    partes.append(f"Cantidad: {int(q) if q.is_integer() else q}")
                except Exception:
                    partes.append(f"Cantidad: {cantidad_val}")
            monto_fmt = _formatear_monto_limpio(monto_export_val or monto_val)
            if monto_fmt != "-":
                partes.append(f"Monto: {monto_fmt}")
        resumen["resumen_aplicador"] = " | ".join(partes) if partes else "-"

    if applier_sin_sku_explicito:
        resumen["resumen_aplicador"] = "ERROR: applier sin SKU explícito"

    if resumen["tipo_promocion"] == "-":
        for x in descuento_items:
            txt = x["msg_plain"]
            m_ambos = re.search(r"Excel\s*\((.*?)\).*?Export\s*\((.*?)\)", txt, re.IGNORECASE)
            if m_ambos:
                resumen["tipo_promocion"] = f"PORCENTUAL - {m_ambos.group(1).strip()}"
                break

    if resumen["resumen_condicion"] == "-":
        for x in condicion_items:
            txt = x["msg_plain"]
            m_lista = re.search(r"misma lista de productos del Excel:\s*\((.*?)\)", txt, re.IGNORECASE)
            if m_lista:
                resumen["resumen_condicion"] = f"Lista: {m_lista.group(1).strip()}"
                break
        if resumen["resumen_condicion"] == "-":
            for x in lista_items:
                txt = x["msg_plain"]
                m = re.search(r"LISTA PRODUCTOS Excel\s*\((.*?)\)", txt, re.IGNORECASE)
                if m:
                    resumen["resumen_condicion"] = f"Lista: {m.group(1).strip()}"
                    break

    lista_locales_resumen = "-"
    locales_resumen = []
    for x in locales_items:
        txt = x["msg_plain"]
        m_lista_local = re.search(r"LISTA LOCAL\s*\((.*?)\)", txt, re.IGNORECASE)
        if m_lista_local:
            lista_locales_resumen = m_lista_local.group(1).strip()
        for m_local in re.finditer(r"LOCAL\s*\((.*?)\)", txt, re.IGNORECASE):
            local = m_local.group(1).strip()
            if local and local not in locales_resumen:
                locales_resumen.append(local)

    if lista_locales_resumen != "-":
        if resumen["resumen_condicion"] == "-":
            resumen["resumen_condicion"] = f"Lista locales: {lista_locales_resumen}"
        elif f"Lista locales: {lista_locales_resumen}" not in resumen["resumen_condicion"]:
            resumen["resumen_condicion"] += f" | Lista locales: {lista_locales_resumen}"
    elif locales_resumen:
        locales_txt = " - ".join(locales_resumen)
        if resumen["resumen_condicion"] == "-":
            resumen["resumen_condicion"] = f"Locales: {locales_txt}"
        elif f"Locales: {locales_txt}" not in resumen["resumen_condicion"]:
            resumen["resumen_condicion"] += f" | Locales: {locales_txt}"

    esperado_dias = None
    export_dias = None
    for x in dias_items:
        txt = x["msg_plain"]
        if "Excel indica restricción" in txt or "Excel indica restriccion" in txt:
            if "L-J" in txt or "lunes y jueves" in txt.lower():
                esperado_dias = "Lunes y Jueves"
        if "Export respeta L-J" in txt:
            m_ok = re.search(r":\s*(.*?)\s*$", txt)
            export_dias = m_ok.group(1).strip() if m_ok else "Lunes, Jueves"
        elif "Export no respeta L-J" in txt:
            m_err = re.search(r"Esperado\s*\((.*?)\)\s*pero trae\s*(.*)$", txt, re.IGNORECASE)
            if m_err:
                esperado_dias = esperado_dias or m_err.group(1).strip()
                export_dias = m_err.group(2).strip()
            else:
                m_trae = re.search(r"trae\s*(.*)$", txt, re.IGNORECASE)
                if m_trae:
                    export_dias = m_trae.group(1).strip()
        elif "No se pudo leer daysAndHours" in txt:
            export_dias = "No se pudo leer Export"

    if dias_items:
        if any(x["tipo"] == "ERR" for x in dias_items):
            resumen["restriccion_dias_estado"] = "No coinciden"
            if esperado_dias and export_dias:
                resumen["restriccion_dias"] = f"ERROR — Esperado {esperado_dias} | Export {export_dias}"
            elif esperado_dias:
                resumen["restriccion_dias"] = f"ERROR — Esperado {esperado_dias}"
            else:
                resumen["restriccion_dias"] = "ERROR — Restricción de días incorrecta"
        elif any(x["tipo"] == "OK" for x in dias_items) or esperado_dias:
            resumen["restriccion_dias_estado"] = "Coinciden"
            resumen["restriccion_dias"] = esperado_dias or export_dias or "Lunes y Jueves"

    hay_err_id = any(x["tipo"] == "ERR" for x in id_items)
    hay_err_fact = any(x["tipo"] == "ERR" for x in fact_items)
    hay_err_cond = any(x["tipo"] == "ERR" for x in condicion_items)
    hay_err_applier = applier_sin_sku_explicito or any(x["tipo"] == "ERR" for x in applier_items)
    hay_err_dias = any(x["tipo"] == "ERR" for x in dias_items)
    solo_ext_fecha_inicio = inicio_tipo == "WARN" and fin_tipo == "OK" and not hay_err_id and not hay_err_fact and not hay_err_cond and not hay_err_applier and not hay_err_dias

    if solo_ext_fecha_inicio:
        resumen["mensaje_principal"] = "Coinciden"
        resumen["aviso_principal"] = "Posible extensión: fecha inicio diferente"
    else:
        if hay_err_id or hay_err_fact or hay_err_cond or hay_err_applier or hay_err_dias or fin_tipo == "ERR" or (inicio_tipo == "ERR" and fin_tipo != "OK"):
            resumen["mensaje_principal"] = "No coinciden"
        elif inicio_tipo == "OK" and fin_tipo == "OK":
            resumen["mensaje_principal"] = "Coinciden"
        else:
            resumen["mensaje_principal"] = "No coinciden"

    return resumen


def extraer_msje_popup_desde_detalles(detalles):
    resultado = {
        "hay": False,
        "id_msje": "",
        "id_padre": "",
        "mensaje": "No hay",
        "salida": "-",
        "texto": "-",
        "resumen_condicion": "-",
        "resumen_aplicador": "-",
        "detalle": [],
    }

    mensajes = []
    for d in detalles:
        if isinstance(d, tuple):
            tipo, msg = d
        else:
            tipo, msg = d.get("tipo"), d.get("msg")
        plain = _strip_html(msg)
        if plain.startswith("[MSJE]") or plain.startswith("[CONDICIÓN]") or plain.startswith("[APPLIER]"):
            mensajes.append({"tipo": tipo, "msg": msg, "plain": plain})

    rel = next((x for x in mensajes if x["plain"].startswith("[MSJE]") and "corresponde a MSJE / POPUP asociado a promoción" in x["plain"]), None)
    if not rel:
        return resultado

    m = re.search(r"ID Lista Locales\s*\((.*?)\).*?promoción\s*\((.*?)\)", rel["plain"], re.IGNORECASE)
    if not m:
        return resultado

    resultado["hay"] = True
    resultado["id_msje"] = m.group(1).strip()
    resultado["id_padre"] = m.group(2).strip()

    cond = next((x for x in mensajes if x["plain"].startswith("[CONDICIÓN]") and ("condición del MSJE" in x["plain"].lower() or "sku de la condición del msje" in x["plain"].lower() or "Lista de condición del MSJE" in x["plain"])), None)
    apl = next((x for x in mensajes if x["plain"].startswith("[MSJE]") and "Texto del mensaje" in x["plain"]), None)
    salida = next((x for x in mensajes if x["plain"].startswith("[MSJE]") and "Salida del mensaje correcta" in x["plain"]), None)
    resumen = next((x for x in mensajes if x["plain"].startswith("[MSJE]") and "Tipo (MSJE)" in x["plain"]), None)

    if cond:
        resultado["resumen_condicion"] = cond["plain"].replace("[CONDICIÓN]", "").strip()
    if salida:
        ms = re.search(r"\((.*?)\)", salida["plain"])
        if ms:
            resultado["salida"] = ms.group(1).strip()
    if apl:
        mt = re.search(r"Texto del mensaje:\s*\((.*?)\)", apl["plain"], re.IGNORECASE)
        if mt:
            resultado["texto"] = mt.group(1).strip()
    if resumen:
        resultado["resumen_aplicador"] = resumen["plain"].replace("[MSJE]", "").strip()

    resultado["mensaje"] = f"MSJE / POPUP asociado a promoción #{resultado['id_padre']}"
    resultado["detalle"] = [{"tipo": x["tipo"], "msg": x["msg"]} for x in mensajes]
    return resultado


def construir_resultado_web(id_geo, excel_origen, export_origen, promo_info, detalles, analisis, es_msje_popup=False, id_padre="", msje_popup=None):
    msje_popup = msje_popup or {}
    msje_popup_hay = bool(msje_popup.get("hay"))
    msje_popup_id = str(msje_popup.get("id_msje") or "")
    msje_popup_id_padre = str(msje_popup.get("id_padre") or id_geo or "")
    msje_popup_mensaje = msje_popup.get("mensaje") or ("No hay" if not msje_popup_hay else "-")
    msje_popup_salida = msje_popup.get("salida") or "-"
    msje_popup_texto = msje_popup.get("texto") or "-"
    msje_popup_resumen = msje_popup.get("resumen_aplicador") or msje_popup.get("resumen") or "-"
    msje_popup_condicion = msje_popup.get("resumen_condicion") or "-"
    msje_popup_fecha_inicio = msje_popup.get("fecha_inicio") or ""
    msje_popup_fecha_fin = msje_popup.get("fecha_fin") or ""

    busqueda_ids = " ".join(
        p for p in [str(id_geo or "").strip(), str(id_padre or "").strip(), msje_popup_id] if p
    ).strip()

    return {
        "id_geo": str(id_geo),
        "mensaje": analisis["mensaje_principal"],
        "aviso_principal": analisis["aviso_principal"],
        "excel_origen": excel_origen,
        "export_origen": export_origen,
        "promo_info": promo_info,
        "detalle": [{"tipo": d[0], "msg": d[1]} if isinstance(d, tuple) else d for d in detalles],
        "estado_id": analisis["estado_id"],
        "estado_facturar": analisis["estado_facturar"],
        "estado_fechas": analisis["estado_fechas"],
        "estado_condicion": analisis["estado_condicion"],
        "estado_applier": analisis["estado_applier"],
        "fecha_inicio_ok": analisis["fecha_inicio_ok"],
        "fecha_fin_ok": analisis["fecha_fin_ok"],
        "tipo_promocion": analisis["tipo_promocion"],
        "resumen_condicion": analisis["resumen_condicion"],
        "resumen_aplicador": analisis["resumen_aplicador"],
        "restriccion_dias": analisis.get("restriccion_dias", "-"),
        "restriccion_dias_estado": analisis.get("restriccion_dias_estado", "No evaluado"),
        "es_msje_popup": es_msje_popup,
        "id_padre": str(id_padre or ""),
        "msje_popup_hay": msje_popup_hay,
        "msje_popup_id": msje_popup_id,
        "msje_popup_id_padre": msje_popup_id_padre,
        "msje_popup_mensaje": msje_popup_mensaje,
        "msje_popup_salida": msje_popup_salida,
        "msje_popup_texto": msje_popup_texto,
        "msje_popup_resumen": msje_popup_resumen,
        "msje_popup_condicion": msje_popup_condicion,
        "msje_popup_fecha_inicio": msje_popup_fecha_inicio,
        "msje_popup_fecha_fin": msje_popup_fecha_fin,
        "busqueda_ids": busqueda_ids,
    }






def _crear_resultado_extra_export(id_geo, promo_info):
    detalle = [{
        "tipo": "ERR",
        "msg": "Promoción presente en Export pero no encontrada en Excel.",
    }]
    analisis = {
        "mensaje_principal": "No coinciden",
        "aviso_principal": "",
        "estado_id": "No coinciden",
        "estado_facturar": "No evaluado",
        "estado_fechas": "No evaluado",
        "estado_condicion": "No evaluado",
        "estado_applier": "No evaluado",
        "fecha_inicio_ok": None,
        "fecha_fin_ok": None,
        "tipo_promocion": promo_info.get("__tipo_descuento", "-") or "-",
        "resumen_condicion": "-",
        "resumen_aplicador": "-",
        "restriccion_dias": "-",
        "restriccion_dias_estado": "No evaluado",
    }
    return construir_resultado_web(
        id_geo=id_geo,
        excel_origen="No solicitado en Excel",
        export_origen=promo_info.get("__export_origen", "-"),
        promo_info=promo_info,
        detalles=detalle,
        analisis=analisis,
    )


def _obtener_ids_solicitados_excel(df_completar_total, df_eventos_usuario):
    ids = set()

    if df_completar_total is not None and not df_completar_total.empty:
        col_id_geo = next((c for c in df_completar_total.columns if "GEOCOM" in str(c).upper()), None)
        if col_id_geo:
            for valor in df_completar_total[col_id_geo].tolist():
                pid = normalizar_local(str(valor).split(".")[0])
                if pid:
                    ids.add(pid)

    if df_eventos_usuario is not None and not getattr(df_eventos_usuario, "empty", True):
        if "ID GEO" in df_eventos_usuario.columns:
            for valor in df_eventos_usuario["ID GEO"].tolist():
                pid = normalizar_local(str(valor).split(".")[0])
                if pid:
                    ids.add(pid)

    return ids

def _copiar_resultados_para_descarga(resultados):
    limpios = []
    for r in resultados:
        copia = dict(r)
        copia["promo_info"] = dict(copia.get("promo_info") or {})
        copia["detalle"] = list(copia.get("detalle") or [])
        limpios.append(copia)
    return limpios


def _estado_reporte_descarga(resultado):
    mensaje = str(resultado.get("mensaje") or "").strip()
    aviso = str(resultado.get("aviso_principal") or "").strip()
    if mensaje == "Coinciden" and aviso:
        return "ATENCION"
    if mensaje == "Coinciden":
        return "OK"
    return "ERROR"


def _valor_si_no(valor):
    return "Sí" if bool(valor) else "No"




def _tipo_descuento_no_wrap(texto):
    txt = str(texto or "-").strip()
    if not txt or txt == "-":
        return "-"
    return txt.replace(" - ", " - ")

def _texto_competencia_prolijo(valor):
    txt = str(valor or "-").strip()
    if not txt or txt == "-":
        return "-"
    txt = txt.replace("Comp. X Producto", "Comp. Por Producto")
    txt = txt.replace("Comp. X Promoción", "Comp. Por Promoción")
    txt = txt.replace("Comp. X Promocion", "Comp. Por Promoción")
    txt = txt.replace("Comp. X Unidades", "Comp. Por Unidades")
    return txt


def _extraer_productos_y_detalle(texto):
    base = str(texto or "-").strip()
    if not base or base == "-":
        return "-", "-"

    partes = [p.strip() for p in base.split("|") if p.strip()]
    if not partes:
        return "-", "-"

    producto = "-"
    detalle = []

    for parte in partes:
        lower = parte.lower()
        if lower.startswith("sku:") or lower.startswith("lista:"):
            if producto == "-":
                producto = parte
            else:
                detalle.append(parte)
        else:
            detalle.append(parte)

    detalle_txt = " | ".join(detalle) if detalle else "-"
    return producto, detalle_txt


def _observacion_reporte(resultado):
    estado = _estado_reporte_descarga(resultado)
    aviso = str(resultado.get("aviso_principal") or "").strip()
    if estado == "OK":
        return "-"
    if aviso:
        return aviso

    detalle = resultado.get("detalle") or []
    preferidos = []
    for d in detalle:
        tipo = str((d or {}).get("tipo") or "").strip().upper()
        msg = _strip_html((d or {}).get("msg", ""))
        if not msg:
            continue
        if tipo == "ERR":
            preferidos.append(msg)
    if preferidos:
        return preferidos[0]

    for d in detalle:
        tipo = str((d or {}).get("tipo") or "").strip().upper()
        msg = _strip_html((d or {}).get("msg", ""))
        if tipo == "WARN" and msg:
            return msg

    return "-"


def _fila_reporte_xlsx(resultado):
    promo_info = resultado.get("promo_info") or {}
    productos_condicion, detalle_condicion = _extraer_productos_y_detalle(resultado.get("resumen_condicion", "-"))
    productos_aplicador, detalle_aplicador = _extraer_productos_y_detalle(resultado.get("resumen_aplicador", "-"))

    fecha_inicio = str(promo_info.get("startDate") or promo_info.get("__start_date") or "-").strip() or "-"
    fecha_fin = str(promo_info.get("endDate") or promo_info.get("__end_date") or "-").strip() or "-"

    return [
        _estado_reporte_descarga(resultado),
        _observacion_reporte(resultado),
        resultado.get("id_geo", "-"),
        fecha_inicio,
        fecha_fin,
        resultado.get("tipo_promocion", "-"),
        promo_info.get("__area_responsable", promo_info.get("area_responsable", "-")),
        _texto_competencia_prolijo(promo_info.get("__tipo_competencia", "-")),
        productos_condicion,
        detalle_condicion,
        productos_aplicador,
        detalle_aplicador,
        promo_info.get("creationUser", "-"),
        resultado.get("excel_origen", "-"),
        resultado.get("export_origen", "-"),
    ]


def _armar_xlsx_resultados(rc, resultados_tradicional, resultados_completar):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    headers = [
        "Estado",
        "Observaciones",
        "ID GEO",
        "Fecha inicio",
        "Fecha fin",
        "Tipo descuento",
        "Área responsable",
        "Tipo competencia",
        "Productos condición",
        "Detalle condición",
        "Productos aplicador",
        "Detalle aplicador",
        "Usuario creador",
        "Excel",
        "Export",
    ]

    title_fill = PatternFill("solid", fgColor="1F4F82")
    title_font = Font(color="FFFFFF", bold=True)
    header_fill = PatternFill("solid", fgColor="DCE6F1")
    header_font = Font(bold=True, color="1F1F1F")
    thin_gray = Side(style="thin", color="D9E0E7")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    ws.merge_cells("A1:O1")
    ws["A1"] = "REPORTE DE VALIDACIÓN DE PROMOCIONES"
    ws["A1"].fill = title_fill
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    ws["A2"] = "Fecha validación"
    ws["B2"] = datetime.now().strftime("%d-%m-%Y")
    ws["D2"] = "Usuario filtrado"
    ws["E2"] = rc or "-"

    for cell in ("A2", "D2"):
        ws[cell].font = Font(bold=True, color="44515E")

    header_row = 4
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_rows = []
    for r in resultados_tradicional:
        data_rows.append(_fila_reporte_xlsx(r))
    for r in resultados_completar:
        data_rows.append(_fila_reporte_xlsx(r))

    start_row = header_row + 1
    for row_idx, row_data in enumerate(data_rows, start=start_row):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        estado = str(row_data[0])
        estado_cell = ws.cell(row=row_idx, column=1)
        estado_cell.font = Font(bold=True)
        if estado == "OK":
            estado_cell.fill = PatternFill("solid", fgColor="DCEFE5")
            estado_cell.font = Font(bold=True, color="115C39")
        elif estado == "ATENCION":
            estado_cell.fill = PatternFill("solid", fgColor="FFF2C9")
            estado_cell.font = Font(bold=True, color="7A5B00")
        else:
            estado_cell.fill = PatternFill("solid", fgColor="F7D9DC")
            estado_cell.font = Font(bold=True, color="8D2430")

    widths = {
        "A": 12, "B": 42, "C": 12, "D": 14, "E": 14,
        "F": 22, "G": 20, "H": 22, "I": 30, "J": 34,
        "K": 30, "L": 34, "M": 18, "N": 38, "O": 26,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:O{max(4, ws.max_row)}"
    ws.sheet_view.showGridLines = False

    for row in range(5, ws.max_row + 1):
        ws.row_dimensions[row].height = 34

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def serializar_resultados(resultados):

    limpios = []
    for r in resultados:
        copia = dict(r)
        copia["promo_info"] = dict(copia.get("promo_info") or {})
        limpios.append(copia)
    return json.dumps(limpios, ensure_ascii=False)


# ============================================================
# RUTA PRINCIPAL
# ============================================================
@app.route("/")
@app.route("/validPromotion/")
def inicio():
    excel, export = listar_archivos()
    return render_template("index.html", excel_files=excel, export_files=export)

# ============================================================
# CONSULTOR DE CARGA
# ============================================================
@app.route("/consultor")
def consultor():
    consultor_backend_ok = callable(construir_consulta)
    return render_template(
        "consultor_ui.html",
        consultor_backend_ok=consultor_backend_ok,
        consultor_catalogos=REGLAS_CONSULTOR.get("catalogos", {}),
        consultor_textos=REGLAS_CONSULTOR.get("textos", {}),
        consultor_requiere_mensaje_tipos=REGLAS_CONSULTOR.get("requiere_mensaje_tipos", []),
    )


@app.route("/consultor/preview", methods=["POST"])
def consultor_preview():
    if not callable(construir_consulta):
        return jsonify({
            "ok": False,
            "error": "El motor del consultor no está disponible.",
        }), 500

    payload = request.get_json(silent=True)
    if not isinstance(payload, dict):
        payload = request.form.to_dict(flat=True)

    try:
        resultado = construir_consulta(payload or {})
        return jsonify({
            "ok": True,
            "resultado": resultado,
        })
    except Exception as e:
        escribir_log(f"ERROR consultor/preview: {e}")
        return jsonify({
            "ok": False,
            "error": str(e),
        }), 400

# ============================================================
# SUBIR ARCHIVOS
# ============================================================
@app.route("/upload", methods=["POST"])
def upload_files():
    cargados_excel = 0
    cargados_export = 0
    export_rechazados = []

    FIRMA_EXPORT = (
        "<?xml version='1.0' encoding='UTF-8'?>"
        "<uy.com.geocom.geopromotion.service.promotion.PromotionBlockList>"
        "<promotionTypeList>"
    )

    for file in request.files.getlist("excel_files"):
        if file and file.filename.lower().endswith(".xlsx"):
            file.save(os.path.join(EXCEL_PATH, file.filename))
            cargados_excel += 1

    for file in request.files.getlist("export_files"):
        if not file or not file.filename.lower().endswith(".txt"):
            continue
        try:
            contenido = file.stream.read(4096)
            file.stream.seek(0)
            texto = None
            for enc in ("utf-8", "utf-16", "latin-1"):
                try:
                    texto = contenido.decode(enc)
                    break
                except UnicodeDecodeError:
                    continue
            if texto is None or FIRMA_EXPORT not in texto.replace("\n", "").replace("\r", ""):
                export_rechazados.append(file.filename)
                continue
            file.save(os.path.join(EXPORT_PATH, file.filename))
            cargados_export += 1
        except Exception:
            export_rechazados.append(file.filename)

    return jsonify({
        "mensaje": "Carga finalizada",
        "excel": cargados_excel,
        "export": cargados_export,
        "excel_cargados": cargados_excel,
        "export_validos": cargados_export,
        "export_rechazados": export_rechazados,
        "lista_excel": os.listdir(EXCEL_PATH),
        "lista_export": os.listdir(EXPORT_PATH),
    })


# ============================================================
# BORRAR ARCHIVOS
# ============================================================
@app.route("/borrar", methods=["POST"])
def borrar_archivos():
    tipo = request.form.get("tipo")
    if tipo == "excel":
        errores = limpiar_carpeta(EXCEL_PATH)
        msg = "Se borraron TODOS los Excel."
    elif tipo == "export":
        errores = limpiar_carpeta(EXPORT_PATH)
        msg = "Se borraron TODOS los Export."
    else:
        return jsonify({"error": "Tipo inválido"})
    return jsonify({
        "mensaje": msg,
        "errores": errores,
        "lista_excel": os.listdir(EXCEL_PATH),
        "lista_export": os.listdir(EXPORT_PATH),
    })


def construir_indices_export(export_files):
    promo_info_por_id = {}
    promos_por_id = {}
    listas_productos_export = {}

    for exp in export_files:
        export_name = os.path.basename(exp)
        tree, raw_text = convertir_txt_a_xml_con_root(exp)
        promos = parsear_promos(tree, export_name=export_name)

        try:
            listas_tmp = parsear_listas_productos_export(exp)
            for nombre, productos in listas_tmp.items():
                listas_productos_export.setdefault(nombre, set()).update(productos)
        except Exception:
            pass

        for promo_dict in promos:
            pid = normalizar_local(str(promo_dict.get("id")).split(".")[0])
            promo_info_por_id[pid] = {
                "creationUser": promo_dict.get("creationUser", "-"),
                "enabled": promo_dict.get("enabled", False),
                "__tipo_competencia": promo_dict.get("__tipo_competencia", "-"),
                "__area_responsable": promo_dict.get("area_name", "-"),
                "__export_origen": promo_dict.get("__export_origen", export_name),
                "__tipo_descuento": "-",
                "startDate": promo_dict.get("startDate", "-"),
                "endDate": promo_dict.get("endDate", "-"),
            }
            if pid not in promos_por_id:
                promos_por_id[pid] = promo_dict

    return promo_info_por_id, promos_por_id, listas_productos_export




def _detectar_columna_id_geocom(df):
    for c in df.columns:
        if "GEOCOM" in str(c).upper():
            return c
    return None


def _detectar_columna_id_lista_cliente(df):
    for c in df.columns:
        nombre = str(c).upper().strip()
        if "ID LISTA CLIENTE" in nombre or nombre == "AG":
            return c
    return None


def _es_excel_club(nombre_archivo):
    return "CLUB" in str(nombre_archivo or "").upper()


def _expandir_filas_club_con_id_lista_cliente(df, nombre_archivo):
    """
    Para archivos CLUB, agrega filas duplicadas usando el valor de
    'ID Lista Cliente' como ID GEO a validar, cuando corresponda.
    Esto permite validar también el clon de convenios sin alterar
    el comportamiento del resto de campañas.
    """
    if df is None or df.empty or not _es_excel_club(nombre_archivo):
        return df

    col_id_geo = _detectar_columna_id_geocom(df)
    col_id_lista_cliente = _detectar_columna_id_lista_cliente(df)

    if not col_id_geo or not col_id_lista_cliente:
        return df

    filas_extra = []

    for _, row in df.iterrows():
        id_geo_base = normalizar_local(str(row.get(col_id_geo)).split(".")[0])
        id_lista_cliente = normalizar_local(str(row.get(col_id_lista_cliente)).split(".")[0])

        if not id_lista_cliente:
            continue

        if id_lista_cliente == id_geo_base:
            continue

        nueva = row.copy()
        nueva[col_id_geo] = id_lista_cliente
        nueva["__id_fuente"] = "ID Lista Cliente"
        filas_extra.append(nueva)

    if not filas_extra:
        return df

    df_base = df.copy()
    if "__id_fuente" not in df_base.columns:
        df_base["__id_fuente"] = "ID Geocom"

    df_extra = df_base.iloc[0:0].copy()
    df_extra = df_extra.reindex(columns=df_base.columns)

    for fila in filas_extra:
        fila_dict = {col: fila[col] if col in fila.index else None for col in df_base.columns}
        df_extra.loc[len(df_extra)] = fila_dict

    return pd.concat([df_base, df_extra], ignore_index=True)


def construir_mapa_area_responsable(excel_files):
    mapa = {}
    for file in excel_files:
        try:
            df_imput = leer_hoja_imput(file)
        except Exception:
            df_imput = None
        if df_imput is None or df_imput.empty:
            continue
        cols = {str(c).strip().upper(): c for c in df_imput.columns}
        col_id = None
        for k, c in cols.items():
            if "GEOCOM" in k or k in {"ID GEO", "ID GEOCOM", "ID"}:
                col_id = c
                break
        col_area = None
        for k, c in cols.items():
            if "AREARESPONSABLE" in k or "AREA RESPONSABLE" in k or k == "AREA":
                col_area = c
                break
        if not col_id or not col_area:
            continue
        for _, row in df_imput.iterrows():
            pid = normalizar_local(row.get(col_id))
            area = normalizar_texto(row.get(col_area))
            if pid and area:
                mapa[pid] = area
    return mapa


# ============================================================
# PROCESAR VALIDACIÓN
# ============================================================
@app.route("/procesar", methods=["POST"])
def procesar():
    rc_web = request.form.get("rc", "").strip().upper()
    excel_files = [os.path.join(EXCEL_PATH, f) for f in os.listdir(EXCEL_PATH) if f.lower().endswith(".xlsx")]
    export_files = [os.path.join(EXPORT_PATH, f) for f in os.listdir(EXPORT_PATH) if f.lower().endswith(".txt")]

    resultados_tradicional = []
    resultados_completar = []
    df_eventos_usuario = None

    promo_info_por_id, promos_por_id, listas_productos_export = construir_indices_export(export_files)
    mapa_area_responsable = construir_mapa_area_responsable(excel_files)

    # FLUJO EVENTOS
    if rc_web:
        try:
            df_eventos_usuario, df_codigos_total, _, archivos_eventos = ejecutar_flujo_eventos(excel_files, rc_externo=rc_web)
        except Exception as e:
            escribir_log(f"ERROR flujo eventos: {e}")
            print(f"ERROR flujo eventos: {e}")
            df_eventos_usuario, df_codigos_total, archivos_eventos = None, pd.DataFrame(), []

        if df_eventos_usuario is not None and not df_eventos_usuario.empty:
            excel_origen_eventos = ", ".join(sorted({os.path.basename(f) for f in archivos_eventos}))
            for id_geo, grupo in df_eventos_usuario.groupby("ID GEO"):
                id_geo_norm = normalizar_local(str(id_geo).split(".")[0])
                promo = promos_por_id.get(id_geo_norm)
                info = promo_info_por_id.get(id_geo_norm, {}).copy()

                if "DESCUENTO" in grupo.columns:
                    val = grupo["DESCUENTO"].iloc[0]
                    if isinstance(val, (int, float)):
                        info["__tipo_descuento"] = _tipo_descuento_no_wrap(f"PORCENTUAL - {int(val * 100) if val <= 1 else int(val)}%")
                    else:
                        info["__tipo_descuento"] = _tipo_descuento_no_wrap(f"PORCENTUAL - {str(val).strip()}")
                else:
                    info["__tipo_descuento"] = "-"

                if promo is None:
                    analisis = {
                        "mensaje_principal": "No existe en export", "aviso_principal": "", "estado_id": "No coinciden",
                        "estado_facturar": "No evaluado", "estado_fechas": "No evaluado", "estado_condicion": "No evaluado",
                        "estado_applier": "No evaluado", "fecha_inicio_ok": None, "fecha_fin_ok": None,
                        "tipo_promocion": "-", "resumen_condicion": "-", "resumen_aplicador": "-",
                    }
                    resultados_tradicional.append(
                        construir_resultado_web(
                            id_geo,
                            excel_origen_eventos,
                            "-",
                            info,
                            [{"tipo": "ERR", "msg": "No encontrada en export"}],
                            analisis,
                        )
                    )
                    continue

                nombre_lista_excel = ""
                if "LISTA PRODUCTOS" in grupo.columns:
                    listas_p = {
                        normalizar_texto(v)
                        for v in grupo["LISTA PRODUCTOS"]
                        if str(v).strip() and normalizar_texto(v)
                    }
                    nombre_lista_excel = next(iter(sorted(listas_p))) if listas_p else ""

                productos_excel = set()
                if nombre_lista_excel:
                    productos_excel = set()

                if not productos_excel and not df_codigos_total.empty and "MARCA" in df_codigos_total.columns:
                    marcas = {
                        normalizar_texto(v)
                        for v in grupo.get("MARCA", [])
                        if str(v).strip() and normalizar_texto(v)
                    }
                    col_sku_codigos = None
                    for candidata in ["CÓDIGO PRODUCTO", "CODIGO PRODUCTO", "SKU"]:
                        if candidata in df_codigos_total.columns:
                            col_sku_codigos = candidata
                            break

                    if col_sku_codigos:
                        for marca in marcas:
                            df_m = df_codigos_total[df_codigos_total["MARCA"].astype(str).str.upper() == marca]
                            productos_excel.update(
                                {
                                    normalizar_texto(x)
                                    for x in df_m[col_sku_codigos].tolist()
                                    if str(x).strip()
                                }
                            )

                _, detalles, msje_popup = validar_promocion_eventos(
                    id_geo,
                    grupo,
                    promo,
                    productos_excel=productos_excel,
                    nombre_lista_excel=nombre_lista_excel,
                    listas_productos_export=listas_productos_export,
                    promos_por_id=promos_por_id,
                )

                analisis = analizar_detalles(detalles)
                n_campana = ""
                for candidata in ["N°CAM", "N° CAM", "N CAM", "N CAMPAÑA", "N°CAMPAÑA", "N° CAMPAÑA"]:
                    if candidata in grupo.columns:
                        valor_campana = grupo[candidata].iloc[0]
                        if str(valor_campana).strip() and str(valor_campana).strip().upper() not in {"NAN", "NONE", "NULL"}:
                            n_campana = str(valor_campana).strip()
                            break
                if n_campana:
                    info["__numero_campana"] = n_campana
                    base_resumen = (analisis.get("resumen_condicion") or "-").strip()
                    analisis["resumen_condicion"] = f"N° campaña: {n_campana}" if base_resumen in {"", "-"} else f"N° campaña: {n_campana} | {base_resumen}"
                info["__tipo_descuento"] = _tipo_descuento_no_wrap(analisis["tipo_promocion"] or info.get("__tipo_descuento", "-"))
                info["__area_responsable"] = analisis.get("area_responsable", info.get("__area_responsable", "-"))
                resultados_tradicional.append(
                    construir_resultado_web(
                        id_geo,
                        excel_origen_eventos,
                        info.get("__export_origen", "-"),
                        info,
                        detalles,
                        analisis,
                        msje_popup=msje_popup,
                    )
                )

    # FLUJO COMPLETAR
    df_completar_total = pd.DataFrame()
    for file in excel_files:
        df_c = leer_hoja_completar(file)
        if df_c is not None and not df_c.empty:
            nombre_excel = os.path.basename(file)
            df_c = _expandir_filas_club_con_id_lista_cliente(df_c, nombre_excel)
            df_c["__excel_origen"] = nombre_excel
            if "__id_fuente" not in df_c.columns:
                df_c["__id_fuente"] = "ID Geocom"
            df_completar_total = pd.concat([df_completar_total, df_c], ignore_index=True)

    if not df_completar_total.empty:
        col_id_geo = [c for c in df_completar_total.columns if "GEOCOM" in str(c).upper()][0]
        for id_geo, grupo in df_completar_total.groupby(col_id_geo):
            id_geo_norm = normalizar_local(str(id_geo).split(".")[0])
            promo = promos_por_id.get(id_geo_norm)
            info = promo_info_por_id.get(id_geo_norm, {}).copy()
            excel_origen = grupo["__excel_origen"].iloc[0]

            if promo is None:
                analisis = {
                    "mensaje_principal": "No existe en export", "aviso_principal": "", "estado_id": "No coinciden",
                    "estado_facturar": "No evaluado", "estado_fechas": "No evaluado", "estado_condicion": "No evaluado",
                    "estado_applier": "No evaluado", "fecha_inicio_ok": None, "fecha_fin_ok": None,
                    "tipo_promocion": "-", "resumen_condicion": "-", "resumen_aplicador": "-",
                }
                resultados_completar.append(construir_resultado_web(id_geo, excel_origen, "-", {}, [{"tipo": "ERR", "msg": "No encontrada en export"}], analisis))
                continue

            _, detalles, msje_popup = validar_promocion_completar(
                id_geo,
                grupo,
                promo,
                listas_productos_export,
                mapa_area_responsable=mapa_area_responsable,
                promos_por_id=promos_por_id,
                retornar_msje_data=True,
            )
            analisis = analizar_detalles(detalles)
            if _es_caso_especial_bycp_3x2(analisis, detalles):
                detalles = _ajustar_detalles_caso_especial_bycp_3x2(detalles)
                analisis = _forzar_analisis_caso_especial_bycp_3x2(analizar_detalles(detalles))
            info["__tipo_descuento"] = _tipo_descuento_no_wrap(analisis["tipo_promocion"] or promo.get("__tipo_descuento", "-"))
            info["__area_responsable"] = analisis.get("area_responsable", info.get("__area_responsable", "-"))

            resultado_principal = construir_resultado_web(
                id_geo,
                excel_origen,
                info.get("__export_origen", "-"),
                info,
                detalles,
                analisis,
                msje_popup=msje_popup,
            )
            resultados_completar.append(resultado_principal)

    ids_solicitados_excel = _obtener_ids_solicitados_excel(df_completar_total, df_eventos_usuario)
    ids_export = set(promo_info_por_id.keys())
    ids_extra_export = sorted(ids_export - ids_solicitados_excel)

    for id_extra in ids_extra_export:
        info_extra = dict(promo_info_por_id.get(id_extra, {}) or {})
        resultados_completar.append(_crear_resultado_extra_export(id_extra, info_extra))

    todos_los_resultados = resultados_tradicional + resultados_completar
    total_promos = len(todos_los_resultados)
    total_ok = sum(1 for r in todos_los_resultados if r.get("mensaje") == "Coinciden" and not r.get("aviso_principal", "").startswith("Posible extensión"))
    total_warn = sum(1 for r in todos_los_resultados if r.get("aviso_principal"))
    total_err = sum(1 for r in todos_los_resultados if r.get("mensaje") != "Coinciden")

    global ULTIMO_REPORTE_DESCARGA
    ULTIMO_REPORTE_DESCARGA = {
        "rc": rc_web,
        "tradicional": _copiar_resultados_para_descarga(resultados_tradicional),
        "completar": _copiar_resultados_para_descarga(resultados_completar),
    }

    return render_template(
        "resultado.html",
        rc=rc_web,
        resultados_tradicional=resultados_tradicional,
        resultados_completar=resultados_completar,
        total_promos=total_promos,
        total_ok=total_ok,
        total_warn=total_warn,
        total_err=total_err,
        tradicional_data=serializar_resultados(resultados_tradicional),
        completar_data=serializar_resultados(resultados_completar),
    )


# ============================================================
# DESCARGAR RESULTADOS
# ============================================================
@app.route("/descargar_resultados", methods=["POST"])
def descargar_resultados():
    global ULTIMO_REPORTE_DESCARGA

    rc = (ULTIMO_REPORTE_DESCARGA or {}).get("rc", "")
    resultados_tradicional = list((ULTIMO_REPORTE_DESCARGA or {}).get("tradicional", []))
    resultados_completar = list((ULTIMO_REPORTE_DESCARGA or {}).get("completar", []))

    if not resultados_tradicional and not resultados_completar:
        salida = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"
        ws["A1"] = "No hay resultados disponibles para descargar."
        ws["A2"] = "Primero ejecuta una validación."
        wb.save(salida)
        salida.seek(0)
        return send_file(
            salida,
            as_attachment=True,
            download_name=f"resultado_validacion_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    salida = _armar_xlsx_resultados(rc, resultados_tradicional, resultados_completar)
    return send_file(
        salida,
        as_attachment=True,
        download_name=f"resultado_validacion_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    app.run(debug=True)
